import openpyxl as xl
import datetime
import pandas as pd
import numpy

def convert_isodate_to_datetime(input_str):
  input_str=input_str[0:len(input_str)-1]
  dateparts=[]
  dateparts.append(str.split(input_str,"T"))
  dateparts.append(str.split(dateparts[0][0],"-"))
  dateparts.append(str.split(dateparts[0][1],":"))
  return_value=datetime.datetime(int(dateparts[1][0]),int(dateparts[1][1]),int(dateparts[1][2]),int(dateparts[2][0]),int(dateparts[2][1]),int(dateparts[2][2]))
  # print(int(dateparts[2][0]),int(dateparts[2][1]),int(dateparts[2][2]))
  return return_value
  

def remove_None_xl(input_list):
  output_list=[]
  for item in  input_list:
    if item.value is not None:
      output_list.append(item)
  return output_list




try:
  workbook=xl.load_workbook(r"C:\Users\rdandekar\Desktop\ProductionProfile.xlsx")
except Exception:
  print('File not found')
  exit()
print(f'Workbook opened \n')

# for worksheet in worksheets:
#   print(worksheet)

# workbook.close()

# profile_date=convert_isodate_to_datetime("2042-01-01T08:15:30Z")
# print(profile_date,type(profile_date))
# pass

# read bulk shift days from excel_input_sheet
first_row=next(workbook['Bulk_Shift_Input'].rows)
days_to_shift=first_row[1].value
print(days_to_shift)
workbook.close()
# k=input("Press any key")

entity_types=['USERDF','SOURCE','SEP','TANK','JOINT','WELL','INLGEN']
entity_dates=[]

worksheets=workbook.sheetnames

df=pd.read_excel(r"C:\Users\rdandekar\Desktop\ProductionProfile.xlsx",sheet_name='SOURCE',header=[0,1],index_col=0,parse_dates=True)
# print(df.columns)
# print(df.columns.levels)
# print(df.loc[slice(None),(slice(None),['CUMGAS','CUMOIL','CUMWAT'])])

req_df=pd.DataFrame(df.loc[slice(None),(slice(None),['CUMGAS','CUMOIL','CUMWAT'])])
del df
# print(req_df)
required_columns=req_df.columns.values
data_values=req_df.values
data_arrays=[[data_values[i,j] for i in range(len(data_values))] for j in range(len(required_columns))]

dates_list=req_df.index.values
days_from_start=[(i-dates_list[0])/numpy.timedelta64(1,'D') for i in dates_list]
new_dates=[datetime.datetime.utcfromtimestamp(i.tolist()/1e9)+datetime.timedelta(days=days_to_shift) for i in dates_list]
dates_list=[datetime.datetime.utcfromtimestamp(i.tolist()/1e9) for i in dates_list]
# print(new_dates)
# print(dates_list)

# for worksheet in worksheets:
#   if worksheet in entity_types:
#     first_column=next(workbook[worksheet].iter_cols(min_col=1,max_col=1))
#     first_column=remove_None_xl(first_column)
#     dates=[]
#     for cell in first_column:
#       dates.append(convert_isodate_to_datetime(str(cell.value)))
#     entity_dates.append([worksheet,dates])
#     merged_cells=workbook[worksheet].merged_cells.ranges
#     for merged_cell in merged_cells:
#         print(merged_cell.cols)
    
  
# for entity_date in entity_dates:
#     print(f'{entity_date}\n\n')
