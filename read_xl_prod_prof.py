import openpyxl as xl
import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
import numpy


def get_new_dates(start_date: datetime.datetime,end_date: datetime.datetime, rollup: str):
    if rollup not in ['monthly','quarterly','semi-annually','annually']:
        print("Incorrect period specification\nShould be one of 'monthly', 'quarterly', 'semi-annually', or 'annually'")
        return []
    if start_date > end_date:
        print("Start date cannot be later than End date\nExiting")
        return []
    dates=[]
    previous_date=start_date
    dates.append(previous_date)

    if rollup=='monthly':
        while True:
            next_date=previous_date+relativedelta(day=31)
            next_date+=relativedelta(days=1)
            next_date=next_date.replace(hour=0,minute=0,second=0,microsecond=0)
            if next_date == end_date:
                dates.append(next_date)
                break
            elif next_date > end_date:
                dates.append(end_date)
                break
            else:
                dates.append(next_date)
                previous_date=next_date
    elif rollup == 'quarterly':
        if start_date.month in [1,2,3]:
            next_date_month=4
            next_date_year=previous_date.year    
        elif start_date.month in [4,5,6]:
            next_date_month=7
            next_date_year=previous_date.year    
        elif start_date.month in [7,8,9]:
            next_date_month=10
            next_date_year=previous_date.year    
        else:
            next_date_month=1
            next_date_year=previous_date.year+1
        next_date=datetime.datetime(next_date_year,next_date_month,1)
        dates.append(next_date)
        previous_date=next_date
        while True:
            next_date=previous_date+relativedelta(months=3)
            if next_date == end_date:
                dates.append(next_date)
                break
            elif next_date > end_date:
                dates.append(end_date)
                break
            else:
                dates.append(next_date)
                previous_date=next_date
    elif rollup=='semi-annually':
        if start_date.month in [1,2,3,4,5,6]:
            next_date_month=7
            next_date_year=previous_date.year   
        else:
            next_date_month=1
            next_date_year=previous_date.year+1
        next_date=datetime.datetime(next_date_year,next_date_month,1)
        dates.append(next_date)
        previous_date=next_date
        while True:
            next_date=previous_date+relativedelta(months=6)
            if next_date == end_date:
                dates.append(next_date)
                break
            elif next_date > end_date:
                dates.append(end_date)
                break
            else:
                dates.append(next_date)
                previous_date=next_date
    elif rollup=='annually':
        while True:
            next_date=datetime.datetime(previous_date.year+1,1,1,0,0,0)
            if next_date == end_date:
                dates.append(next_date)
                break
            elif next_date > end_date:
                dates.append(end_date)
                break
            else:
                dates.append(next_date)
                previous_date=next_date
    return dates      


def get_required_columns(input_df: pd.DataFrame):
  # df=pd.read_excel(workbook,sheet_name=worksheet,header=[0,1],index_col=0,parse_dates=True,engine='openpyxl')
  columns_to_delete=[]
  for tup in input_df.columns.values:
    if "CUM" not in tup[1]:
      columns_to_delete.append(tup)
  for col in columns_to_delete:
    input_df.pop(col)
  return input_df
    
def dates_bulk_shifted(input_dataframe,days_to_shift):

  required_columns=input_dataframe.columns.values
  data_values=input_dataframe.values
  data_arrays=[[data_values[i,j] for i in range(len(data_values))] for j in range(len(required_columns))]
  dates_list=input_dataframe.index.to_numpy()
  dates_list = [datetime.datetime.fromisoformat(x[0:len(x)-1]) for x in dates_list]
  days_from_start=[(i-dates_list[0]).days for i in dates_list]

  new_dates=[j+datetime.timedelta(days=days_to_shift) for j in dates_list]
  new_days_from_start=[(i-new_dates[0]).days for i in new_dates]

  sers = [pd.Series(numpy.interp(new_days_from_start,days_from_start,data_arrays[i])) for i in range(len(required_columns))]
  cols_with_rates,sers_with_rates=get_rates_from_cum(sers,required_columns,new_days_from_start)

  output_df=pd.DataFrame(pd.concat(sers_with_rates,axis=1))
  output_df.index=new_dates
  output_df.columns=pd.MultiIndex.from_tuples(cols_with_rates)

  return output_df

def get_rates_from_cum(series,columns,days_from_start):
  new_columns=[]
  for i in range(len(columns)):
    new_columns.append(columns[i])
    rate_type="AVG._"+columns[i][1].replace('CUM','')+"_RATE"
    new_columns.append((columns[i][0],rate_type))
  
  new_series=[]
  for ser in series:
    new_series.append(ser)
    ser=ser.to_list()
    rate_series=[]
    for i in range(len(ser)-1):
      rate_series.append((ser[i+1]-ser[i])/(days_from_start[i+1]-days_from_start[i]))
    rate_series.append(0)
    rate_series=pd.Series(rate_series)
    new_series.append(rate_series)
  
  return new_columns,new_series

def convert_isodate_to_datetime(input_str):
  input_str=input_str[0:len(input_str)-1]
  dateparts=[]
  dateparts.append(str.split(input_str,"T"))
  dateparts.append(str.split(dateparts[0][0],"-"))
  dateparts.append(str.split(dateparts[0][1],":"))
  return_value=datetime.datetime(int(dateparts[1][0]),int(dateparts[1][1]),int(dateparts[1][2]),int(dateparts[2][0]),int(dateparts[2][1]),int(dateparts[2][2]))
  # print(int(dateparts[2][0]),int(dateparts[2][1]),int(dateparts[2][2]))
  return return_value

def get_shift_duration(workbook):
  wbk=xl.load_workbook(workbook)
  worksheets=wbk.sheetnames
  if 'Bulk_Shift_Input' in worksheets:
    first_row=next(wbk['Bulk_Shift_Input'].rows)
    wbk.close()
    del wbk
    return(first_row[1].value)
  else:
    wbk.close()
    del wbk
    return None

    
def dates_bulk_shifted_rolledup(input_dataframe,days_to_shift,rollup):

  required_columns=input_dataframe.columns.values
  data_values=input_dataframe.values
  data_arrays=[[data_values[i,j] for i in range(len(data_values))] for j in range(len(required_columns))]
  dates_list=input_dataframe.index.to_numpy()
  dates_list = [datetime.datetime.fromisoformat(x[0:len(x)-1]) for x in dates_list]
  days_from_start=[(i-dates_list[0]).days for i in dates_list]
  new_start_date=dates_list[0]+datetime.timedelta(days=days_to_shift)
  new_end_date=dates_list[-1]+datetime.timedelta(days=days_to_shift)
  if rollup=="None":
    new_dates=[j+datetime.timedelta(days=days_to_shift) for j in dates_list]
    # new_days_from_start=[(i-new_dates[0]).days for i in new_dates]
  elif rollup=="monthly":
    new_dates=get_new_dates(new_start_date,new_end_date,'monthly')
  elif rollup=="quarterly":
    new_dates=get_new_dates(new_start_date,new_end_date,'quarterly')
  elif rollup=="semi-annually":
    new_dates=get_new_dates(new_start_date,new_end_date,'semi-annually')
  elif rollup=='annually':
    new_dates=get_new_dates(new_start_date, new_end_date,'annually')
  else:
    new_dates=dates_list
  new_days_from_start=[(i-new_dates[0]).days for i in new_dates]
  sers = [pd.Series(numpy.interp(new_days_from_start,days_from_start,data_arrays[i])) for i in range(len(required_columns))]
  cols_with_rates,sers_with_rates=get_rates_from_cum(sers,required_columns,new_days_from_start)

  output_df=pd.DataFrame(pd.concat(sers_with_rates,axis=1))
  output_df.index=new_dates
  output_df.columns=pd.MultiIndex.from_tuples(cols_with_rates)

  return output_df


print("Hello World")
shift_duration=get_shift_duration(r"C:\Users\rdandekar\Desktop\Prod_Prof.xlsx")
if shift_duration is None:
  exit()
entity_types=['USERDF','SOURCE','SEP','TANK','JOINT','WELL','INLGEN']
df_dict={}
print("\nLoading workbook...\n")
wbk=xl.load_workbook(r"C:\Users\rdandekar\Desktop\Prod_Prof.xlsx")
for entity_type in entity_types:
  print(f"Starting to process {entity_type} sheet\n")
  # input_df=pd.DataFrame()
  # df=pd.read_excel(wbk,sheet_name=entity_type,header=[0,1],index_col=0,parse_dates=True,engine='openpyxl')
  input_df=pd.read_excel(wbk,sheet_name=entity_type,header=[0,1],index_col=0,engine='openpyxl')
  input_dates=input_df.index.to_numpy()
  input_df.index=input_dates
  input_df=get_required_columns(input_df)
  
  if not input_df.empty:
    # df=dates_bulk_shifted(df,shift_duration)
    df=dates_bulk_shifted_rolledup(input_df,shift_duration,'None')
    df_dict[entity_type+"_SHIFTED"]=df
    df=dates_bulk_shifted_rolledup(input_df,shift_duration,'monthly')
    df_dict[entity_type+"_SHIFTED_MONTHLY"]=df
    df=dates_bulk_shifted_rolledup(input_df,shift_duration,'quarterly')
    df_dict[entity_type+"_SHIFTED_QUARTERLY"]=df
    df=dates_bulk_shifted_rolledup(input_df,shift_duration,'annually')
    df_dict[entity_type+"_SHIFTED_ANNUALLY"]=df
  print(f"Finished processing {entity_type} sheet\n")

print("\nWriting to workbook...\n\n")
writer = pd.ExcelWriter(r"C:\Users\rdandekar\Desktop\Prod_Prof_shifted.xlsx", engine='openpyxl')
for key in df_dict:
  print(f'Writing {key} sheet...\n')
  df_dict[key].to_excel(writer,sheet_name=key)
print('Finished writing all sheets\nSaving and closing file...\n')
writer.save()
writer.close()
print('Closed file, exiting!\n')
exit()




